# -*- coding: utf-8 -*-

import os
import datetime

import psycopg2
import psycopg2.extras
import DialogCompareQuery

from PyQt5 import QtCore, QtGui, QtWidgets, uic
from PyQt5.QtWidgets import (QMainWindow, QApplication, QTableWidgetItem, QDialog, QFileDialog, QMessageBox, QAbstractItemView)

from openpyxl import (load_workbook, Workbook)
from utils import (dbconn, commfunc)

ui_folder = os.path.abspath(os.path.dirname('__ui__/'))
form_class = uic.loadUiType(os.path.join(ui_folder, "SqlConversion.ui"))[0]

class MyWindow(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        #체크된 row
        self.checkedRows = set([])

        #button 클릭 정의
        self.btnSqlMngSearch.clicked.connect(self.searchSql)
        self.btnSaveSqlInsertExcel.clicked.connect(self.saveSqlInsertExcel)
        self.btnDownSqlUploadExcelSample.clicked.connect(self.downSqlUploadExcelSample)
        self.btnDownSqlListExcel.clicked.connect(self.downSqlListExcel)
        self.btnActSQLConversion.clicked.connect(self.doSqlConversion)

        #TableWidget 클릭 연결
        self.tblwSqlMngSqlList.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tblwSqlMngSqlList.itemClicked.connect(self.handleItemClicked)
        self.tblwSqlMngSqlList.doubleClicked.connect(self.handleItemDoubleClicked)

    def getSqlList(self, listType=1):
        valQryCl = self.cbSqlMngQryCl.currentText()
        valSqlId = self.leSqlMngSqlId.text()
        valConYn = self.cbSqlMngConYn.currentText()
        valConDate = self.cbSqlMngConDate.currentText()

        if valQryCl == "전체": valQryCl = "";
        if valConYn == "전체": valConYn = "";
        if valConDate == "전체": valConDate = "";

        try:
            con = dbconn.createConnection('PGS')
            cur = con.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            if listType == 1:
                qrystr = """SELECT 
                                A.file_nm, A.sql_id, A.job_cl, A.qry_cl, 
                                (CASE A.wk_stat_cd WHEN 'C' THEN 'Y' ELSE 'N' END) AS convert_yn,
                                TO_CHAR(A.rgs_dttm, 'YYYY.MM.DD HH24:MI') as rgs_dttm
                            FROM
                                B2EN_SC_SQL_LIST A
                            where 1=1"""
            elif listType == 2:
                qrystr = """SELECT 
                                A.file_nm, A.sql_id, A.job_cl, A.qry_cl, 
                                (CASE A.wk_stat_cd WHEN 'C' THEN 'Y' ELSE 'N' END) AS convert_yn,
                                string_agg(B.sql_text, chr(10) order by B.line asc) as asis_sql,
                                string_agg(C.sql_text, chr(10) order by C.line asc) as tobe_sql,
                                TO_CHAR(A.rgs_dttm, 'YYYY.MM.DD HH24:MI') as rgs_dttm
                            FROM
                                B2EN_SC_SQL_LIST A
                                INNER JOIN B2EN_SC_SQL_TEXT B ON A.file_nm = B.file_nm AND A.sql_id = B.sql_id
                                LEFT OUTER JOIN B2EN_SC_SQL_TEXT_RST C on B.file_nm = C.file_nm and B.sql_id = C.sql_id and B.line = C.line
                            where 1=1"""

            if valQryCl != "":
                qrystr += " and A.qry_cl = '{0}'".format(valQryCl);

            if valSqlId != "":
                qrystr += " and A.sql_id = '{0}'".format(valSqlId);

            if valConYn == "Y":
                qrystr += " and A.wk_stat_cd = 'C'";
            elif valConYn == "N":
                qrystr += " and A.wk_stat_cd != 'C'";

            if valConDate != "":
                # 어떤 조건이 포함되어야 하는지 확인 필요
                qrystr += " and 1=1";

            if listType == 1:
                qrystr += " ORDER BY A.rgs_dttm DESC"

            if listType == 2:
                qrystr += " GROUP BY A.file_nm, A.sql_id, A.job_cl, A.qry_cl, A.wk_stat_cd, A.rgs_dttm"

            print(qrystr)

            cur.execute(qrystr)
            rows = cur.fetchall()
        except Exception as e:
            print("Exception 발생 : ", e);
            sys.exit(app.exec())
        finally:
            con.close()
            con = None

        return rows

    def searchSql(self):
        tblSqlList = self.tblwSqlMngSqlList
        tblSqlList.setRowCount(0)

        rows = self.getSqlList(1)

        # itemList = [ \
        #     {'fileName': 'selx.xml', 'sqlId': 'selx.selUserList', 'sqlCl': 'SELECT', 'conversionYN': 'Y',
        #      'conversionDate': '2017.06.23 14:30'}, \
        #     {'fileName': 'selx.xml', 'sqlId': 'selx.insUser', 'sqlCl': 'INSERT', 'conversionYN': 'N',
        #      'conversionDate': ''} \
        #     ];

        # itemList를 filtering할 경우
        # if valSqlCl != "":
        #     itemList = [item for item in itemList if item.get("sqlcl") == valSqlCl]
        # if valSqlId != "":
        #     itemList = [item for item in itemList if item.get("sqlid") == valSqlId]
        tblSqlList.setRowCount(len(rows))
        rownum = 0
        for row in rows:
            item_0 = QTableWidgetItem("")
            item_0.setFlags(QtCore.Qt.ItemIsUserCheckable |
                            QtCore.Qt.ItemIsEnabled)
            item_0.setCheckState(QtCore.Qt.Unchecked)

            tblSqlList.setItem(rownum, 0, item_0)
            tblSqlList.setItem(rownum, 1, QTableWidgetItem(row.get("file_nm")))
            tblSqlList.setItem(rownum, 2, QTableWidgetItem(row.get("sql_id")))
            tblSqlList.setItem(rownum, 3, QTableWidgetItem(row.get("qry_cl")))
            tblSqlList.setItem(rownum, 4, QTableWidgetItem(row.get("convert_yn")))
            tblSqlList.setItem(rownum, 5, QTableWidgetItem(row.get("rgs_dttm")))

            rownum += 1

        tblSqlList.resizeColumnsToContents()

        # Edit하지 않도록 수정(default로 수정 가능함)
        tblSqlList.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        # tblSqlList.resizeRowsToContents()

        # 체크된 row-list 초기화
        self.checkedRows.clear()

    def handleItemClicked(self, item):
        if item.checkState() == QtCore.Qt.Checked:
            # print("체크되었네요", item.row())
            commfunc.setTableWidgetRowColor(self.tblwSqlMngSqlList, item.row(), QtGui.QColor(85,170,255))
            self.checkedRows.add(item.row())
        else:
            # print("체크되지 않았네요")
            commfunc.setTableWidgetRowColor(self.tblwSqlMngSqlList, item.row(), QtGui.QColor(255, 255, 255))
            if item.row() in self.checkedRows: self.checkedRows.remove(item.row());

        file_nm = self.tblwSqlMngSqlList.item(item.row(), 1).text()
        sql_id  = self.tblwSqlMngSqlList.item(item.row(), 2).text()

        try:
            con = dbconn.createConnection('PGS')
            cur = con.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            qrystr = """select 
                        A.file_nm, A.sql_id, 
                        string_agg(A.sql_text, chr(10) order by A.line asc) as asis_sql,
                        string_agg(B.sql_text, chr(10) order by B.line asc) as tobe_sql
                    from
                        B2EN_SC_SQL_TEXT A
                        left outer join B2EN_SC_SQL_TEXT_RST B on A.file_nm = B.file_nm and A.sql_id = B.sql_id and A.line = B.line
                    where
                        A.file_nm = '{0}' and A.sql_id = '{1}'
                    group by A.file_nm, A.sql_id""".format(file_nm, sql_id)
            print(qrystr)

            cur.execute(qrystr)
            rows = cur.fetchall()

            self.txbSqlMngAsisSQL.setText(rows[0].get("asis_sql"))
            self.txbSqlMngTobeSQL.setText(rows[0].get("tobe_sql"))
        except Exception as e:
            print("Exception 발생 : ", e);
            sys.exit(app.exec())
        finally:
            con.close()
            con = None

    def handleItemDoubleClicked(self, item):
        #print(item.row(),"가 더블클릭 되었습니다")
        file_nm = self.tblwSqlMngSqlList.item(item.row(), 1).text()
        sql_id = self.tblwSqlMngSqlList.item(item.row(), 2).text()
        conversionYn = self.tblwSqlMngSqlList.item(item.row(), 4).text()

        if conversionYn == "N":
            QMessageBox.warning(self, 'No Conversioned SQL', "SQL 비교분석은 컨버전 완료된 SQL만 가능합니다.",
                                QMessageBox.Ok, QMessageBox.Ok)
        else:
            compareQuery = DialogCompareQuery.CompareQueryWindow()
            compareQuery.viewCompareQuery(file_nm, sql_id)
            compareQuery.show()
            compareQuery.exec_()

    #SQL 엑셀 업로드
    def saveSqlInsertExcel(self):
        fname = QFileDialog.getOpenFileName(self, '엑셀파일 선택', '', '');
        boolInsertSuccess = False

        if fname[0]:
            wb = load_workbook(filename=fname[0])
            colDict = [\
                            {'colName': 'fileNm', 'colMandatoryYN': 'Y', 'colMaxLength': '100'}, \
                            {'colName': 'sqlId',  'colMandatoryYN': 'Y', 'colMaxLength': '100'}, \
                            {'colName': 'jobCl',  'colMandatoryYN': 'N', 'colMaxLength': '1'}, \
                            {'colName': 'sqlText', 'colMandatoryYN': 'Y', 'colMaxLength': '100'} \
                ]
            excelRows = commfunc.getExcelData(wb, colDict)

            sqlList = []  # sql-list 데이터
            sqlTextList = []  # sqlText의 리스트(쿼리별)
            sqlFullList = []  # sql + 실제쿼리 데이터
            lineNum = 1  # 각 쿼리별 line
            sqlidx = -1  # sql-list index

            # 엑셀 ROW를 돌면서 data-set 만들기
            for excelRow in excelRows:
                ##해당 ROW가 쿼리의 첫 라인일 경우
                if {'fileNm': excelRow['fileNm'], 'sqlId': excelRow['sqlId'], 'jobCl': excelRow['jobCl']} not in sqlList:
                    sqlidx += 1
                    sqlTextList = []
                    sqlList.append({'fileNm': excelRow['fileNm'], 'sqlId': excelRow['sqlId'], 'jobCl': excelRow['jobCl']})
                    sqlFullList.append({'fileNm': excelRow['fileNm'], 'sqlId': excelRow['sqlId'], 'jobCl': excelRow['jobCl'],'sqlTextList': sqlTextList})

                sqlTextList.append(excelRow['sqlText'])
                sqlFullList[sqlidx]["sqlTextList"] = sqlTextList

            try:
                con = dbconn.createConnection('PGS')
                cur = con.cursor()

                for sqlRec in sqlFullList:
                    fileNm      = sqlRec['fileNm']
                    sqlId       = sqlRec['sqlId']
                    jobCl       = sqlRec['jobCl']
                    sqlTextList = sqlRec['sqlTextList']
                    sqlFullText = "\n".join(sqlTextList)
                    qryCl       = commfunc.getQueryType(sqlFullText)

                    ## B2EN_SC_SQL_LIST에서 해당 쿼리 삭제
                    cur.execute("DELETE FROM B2EN_SC_SQL_LIST WHERE file_nm = %s and sql_id = %s", (fileNm, sqlId))

                    ## B2EN_SC_SQL_TEXT에서 해당 쿼리의 전체 데이터 삭제
                    cur.execute("DELETE FROM B2EN_SC_SQL_TEXT WHERE file_nm = %s and sql_id = %s", (fileNm, sqlId))

                    qryFullStr = ""
                    ## sqlTextList를 돌면서
                    for qryLineStr in sqlRec["sqlTextList"]:
                        if qryFullStr == "":
                            qryFullStr +=  qryLineStr;
                        else :
                            qryFullStr += "\n" + qryLineStr;

                    ## B2EN_SC_SQL_LIST 테이블에 입력
                    cur.execute("insert into B2EN_SC_SQL_LIST (file_nm, sql_id, job_cl, qry_cl, wk_stat_cd, rgs_dttm) values (%s, %s, %s, %s, 'R', now())", \
                             (fileNm, sqlId, jobCl, qryCl))

                    lineNum = 1
                    for qryLineStr in sqlTextList:
                        ### B2EN_SC_SQL_TEXT에 ROW-DATA 입력
                        cur.execute("insert into B2EN_SC_SQL_TEXT (file_nm, sql_id, line, sql_text) values (%s, %s, %s, %s)", \
                                 (fileNm, sqlId, lineNum, qryLineStr))

                        lineNum += 1
                con.commit()
                boolInsertSuccess = True
            except Exception as e:
                print("Exception 발생 : ", e);
                sys.exit(app.exec())
            finally:
                con.close()
                con = None

        if boolInsertSuccess == True:
            QMessageBox.information(self, 'SQL Insert Success', "SQL 파일 업로드가 완료되었습니다.",
                                               QMessageBox.Ok, QMessageBox.Ok)
            self.cbSqlMngQryCl.setCurrentIndex(0)
            self.leSqlMngSqlId.setText("")
            self.cbSqlMngConYn.setCurrentIndex(2)
            self.cbSqlMngConDate.setCurrentIndex(0)

            self.searchSql()

    def downSqlUploadExcelSample(self):
        bufSize = 1024

        fileName = QFileDialog.getSaveFileName(self, "saveFile", "./SQLUploadSample.xlsx")
        print(fileName[0])
        #출력파일 체크
        if fileName[0] != '':
            f = open('SQLUploadSample.xlsx', 'rb')
            h = open(fileName[0], 'wb')
            data = f.read(bufSize)
            while data:
                h.write(data)
                data = f.read(bufSize)

            f.close()
            h.close()

    def downSqlListExcel(self):
        rows = self.getSqlList(2)
        now = datetime.datetime.now()

        fileName = QFileDialog.getSaveFileName(self, "saveFile", "./SQLData_"+now.strftime('%Y%m%d')+".xlsx")

        if fileName[0] != '':
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "SQL data"
            ws1["A1"].value = "파일명"
            ws1["B1"].value = "SQL ID"
            ws1["C1"].value = "업무구분"
            ws1["D1"].value = "쿼리구분"
            ws1["E1"].value = "컨버전여부"
            ws1["F1"].value = "ASIS SQL"
            ws1["G1"].value = "TOBE SQL"
            ws1["H1"].value = "등록일시"

            i = 2
            for row in rows:
                ws1["A" + str(i)].value = row.get("file_nm")
                ws1["B" + str(i)].value = row.get("sql_id")
                ws1["C" + str(i)].value = row.get("job_cl")
                ws1["D" + str(i)].value = row.get("qry_cl")
                ws1["E" + str(i)].value = row.get("convert_yn")
                ws1["F" + str(i)].value = row.get("asis_sql")
                ws1["G" + str(i)].value = row.get("tobe_sql")
                ws1["H" + str(i)].value = row.get("rgs_dttm")

                i += 1

            wb.save(filename=fileName[0])

    def doSqlConversion(self):
        if len(self.checkedRows) == 0:
            QMessageBox.warning(self, 'Convert Failure', "선택된 SQL이 없습니다.\nSQL을 선택해주십시오",
                                    QMessageBox.Ok, QMessageBox.Ok)
        else:
            for checkedRow in self.checkedRows:
                file_nm = self.tblwSqlMngSqlList.item(checkedRow, 1).text()
                sql_id = self.tblwSqlMngSqlList.item(checkedRow, 2).text()

                print(file_nm, ":", sql_id)


if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    mywin = MyWindow()
    mywin.show()
    sys.exit(app.exec())