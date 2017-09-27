# -*- coding: utf-8 -*-
# 작성자 : 김근호
# 프로그램 : 컬럼 매핑 프로그램

import sys
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QtCore import *
from DBaseClass import DBase
from openpyxl import load_workbook

form_class = uic.loadUiType("ColumnMapping.ui")[0]

class MyWindow(QMainWindow, form_class, DBase):
    def __init__(self):
        dbname = "postgres"
        user = "postgres"
        host = "localhost"
        passwd = ''

        super().__init__()
        #print('dbname=%s user=%s host=%s password=%s' % (dbname, user, host, passwd))
        DBase.__init__(self, dbname, user, host, passwd)

        self.setupUi(self)
        self.cur = DBase.cursor(self)

        self.tableWidget.setColumnCount(9)
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)

        self.pushSearchButton.clicked.connect(self.setTableWidgetData)
        self.pushInputButton.clicked.connect(self.ButtonInputClicked)
        self.pushSampleButton.clicked.connect(self.ButtonSampleClicked)

        self.setTableWidgetData()

    def setTableWidgetData(self):
        column_headers = ['ASIS논리명','ASIS테이블명','ASIS논리속성명','ASIS컬럼명','TOBE논리명','TOBE테이블명','TOBE논리속성명','TOBE컬럼명','등록일시']
        self.tableWidget.setHorizontalHeaderLabels(column_headers)
        lineASISTableValue = self.lineASISTableEdit.text()
        lineASISColumnValue = self.lineASISColumnEdit.text()
        lineTOBETableValue = self.lineTOBETableEdit.text()
        lineTOBEColumnValue = self.lineTOBEColumnEdit.text()

        sql_string = "SELECT asis_logical_tab" \
                     "      ,asis_tab" \
                     "      ,asis_logical_col" \
                     "      ,asis_col" \
                     "      ,tobe_logical_tab" \
                     "      ,tobe_tab" \
                     "      ,tobe_logical_col" \
                     "      ,tobe_col" \
                     "      ,to_char(rgs_dttm,'YYYYMMDD HH24MISS') rgs_dttm " \
                     "FROM B2EN_SC_COL_MAP " \
                     "WHERE 1=1"
        if lineASISTableValue  != "": sql_string += " and (asis_tab like '%{0}%' or asis_logical_tab like '%{0}%')".format(lineASISTableValue)
        if lineASISColumnValue != "": sql_string += " and (asis_col like '%{0}%' or asis_logical_col like '%{0}%')".format(lineASISColumnValue)
        if lineTOBETableValue  != "": sql_string += " and (tobe_tab like '%{0}%' or tobe_logical_tab like '%{0}%')".format(lineTOBETableValue)
        if lineTOBEColumnValue != "": sql_string += " and (tobe_col like '%{0}%' or tobe_logical_col like '%{0}%')".format(lineTOBEColumnValue)
        sql_string += ' ORDER BY rgs_dttm DESC NULLS LAST;'

        print(sql_string)
        self.cur.execute(sql_string)
        #검색전 데이터 초기화
        self.tableWidget.setRowCount(0)

        rownum = 0
        for row in self.cur:
            self.tableWidget.insertRow(rownum)
            self.tableWidget.setItem(rownum, 0, QTableWidgetItem(row[0]))
            self.tableWidget.setItem(rownum, 1, QTableWidgetItem(row[1]))
            self.tableWidget.setItem(rownum, 2, QTableWidgetItem(row[2]))
            self.tableWidget.setItem(rownum, 3, QTableWidgetItem(row[3]))
            self.tableWidget.setItem(rownum, 4, QTableWidgetItem(row[4]))
            self.tableWidget.setItem(rownum, 5, QTableWidgetItem(row[5]))
            self.tableWidget.setItem(rownum, 6, QTableWidgetItem(row[6]))
            self.tableWidget.setItem(rownum, 7, QTableWidgetItem(row[7]))
            self.tableWidget.setItem(rownum, 8, QTableWidgetItem(row[8]))
            rownum = rownum+1

        #테이블 사이즈를 값에 맞추기
        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.resizeRowsToContents()

    def ButtonInputClicked(self):
        pass
        rno = 1
        cno = 1
        fname = QFileDialog.getOpenFileName(self,"inputFile")
        print(fname)
        #input file check
        if fname[0] != '':
            wb = load_workbook(fname[0])
            work_sheet = wb.get_sheet_by_name('Example1')

            for row in work_sheet.iter_rows(row_offset=1):
                #print(row)
                insert_row = dict()
                for cell in row:
                    #print(cell)
                    if cell.value is not None:
                        insert_row[cno] = cell.value
                        #print(insert_row[cno])
                    cno = cno + 1
                #print(insert_row)
                #빈줄 제거

                if 1 in insert_row:
                    #print(rno,len(insert_row), "행 ", insert_row[1], insert_row[2], insert_row[3], insert_row[4])
                    if len(insert_row) == 8:
                        #sql_string = "insert into B2EN_SC_TAB_MAP values('%s','%s',%d);" % (insert_row[1], insert_row[2], insert_row[3])
                        sql_string = """with upsert as(
                                           update B2EN_SC_COL_MAP set asis_logical_tab=UPPER('%s'), asis_logical_col=UPPER('%s'),
                                                                      tobe_logical_tab=UPPER('%s'), tobe_logical_col=UPPER('%s'),
                                                                      rgs_dttm=current_timestamp
                                           where UPPER(asis_tab)=UPPER('%s')
                                           and UPPER(asis_col)=UPPER('%s')
                                           and UPPER(tobe_tab)=UPPER('%s')
                                           and UPPER(tobe_col)=UPPER('%s')
                                           returning *
                                    )
                                    insert into B2EN_SC_COL_MAP(asis_logical_tab,asis_tab,asis_logical_col,asis_col,tobe_logical_tab,tobe_tab,tobe_logical_col,tobe_col,rgs_dttm)
                                    select '%s', UPPER('%s'), '%s', UPPER('%s'), '%s', UPPER('%s'), '%s', UPPER('%s'), current_timestamp
                                    where not exists(
                                                     select UPPER(asis_tab), UPPER(tobe_tab)
                                                     from upsert
                                                     where UPPER(asis_tab)=UPPER('%s')
                                                     and UPPER(asis_col)=UPPER('%s')
                                                     and UPPER(tobe_tab)=UPPER('%s')
                                                     and UPPER(tobe_col)=UPPER('%s'));
                                 """ % (insert_row[1],insert_row[3],insert_row[5],insert_row[7],
                                              insert_row[2],insert_row[4],insert_row[6],insert_row[8],
                                              insert_row[1],insert_row[2],insert_row[3],insert_row[4],insert_row[5],insert_row[6],insert_row[7],insert_row[8],
                                              insert_row[2], insert_row[4], insert_row[6], insert_row[8])
                        print(sql_string)
                        try:
                            self.cur.execute(sql_string)
                            DBase.commit(self)
                        except Exception as e:
                            print("insert error:", e)
                            DBase.rollback(self)
                    else:
                        QMessageBox.critical(None, "Message", str(rno)+"행의 입력값에 오류가 발생했습니다.",QMessageBox.Cancel)

                rno = rno + 1
                cno = 1

            self.setTableWidgetData()

    def ButtonSampleClicked(self):
        bufsize = 1024

        filename = QFileDialog.getSaveFileName(self, "saveFile", "C:/ColumnMappingSample.xlsx")
        print(filename[0])
        #출력파일 체크
        if filename[0] != '':
            f = open('ColumnMappingSample.xlsx', 'rb')
            h = open(filename[0], 'wb')

            data = f.read(bufsize)
            while data:
                h.write(data)
                data = f.read(bufsize)

            f.close()
            h.close()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = MyWindow()
    myWindow.show()
    app.exec_()
